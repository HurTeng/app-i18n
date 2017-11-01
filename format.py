#!/usr/bin/python2.7
# coding=utf-8
#
# format_txt.py
#
# Author: HurTeng
# Created: September 25, 2017

import sys
import os
import re
import codecs
import shutil
from xml.dom.minidom import Document
from openpyxl import load_workbook

TEXT_FILE_DIR_PREFIX = "values-"

class unicodetxt_to_formattxt_obj:
    # 初始化
    def __init__(self):
        # if len(sys.argv) != 2:
        #     print 'usage: ./txt2xml.py file'
        #     sys.exit(1)

        self.txtfd = 0  # 文本
        self.langls = []  # 语言的list
        self.infols = []  # 信息的list

    def change_file(self, xlsx_fname):
        wb = load_workbook(xlsx_fname)
        sheets = wb.sheetnames
        for sh in sheets:
            print(sh)  # ['Sheet1', 'Sheet2', 'Sheet3']
        sheet = wb.get_sheet_by_name(sheets[0])
        print(sheet)
        # print(sheet["C"])  # 第C列

        # 获取语言列表
        langs = []  # 第1行
        for l in sheet["1"][1:]:
            if l.value is not None:
                print(l.column)
                print(l.value)
                langs.append({"key": l.column, "value": l.value})

        index_ls = sheet["A"]
        for l in langs:
            col = l["key"]
            name = l["value"]
            strs = sheet[str(col)]
            stringls = []
            for n in range(len(strs)):
                k = index_ls[n].value
                v = strs[n].value

                if k:
                    k = k.encode("utf-8")
                else:
                    k = ""
                if v:
                    v = v.encode("utf-8")
                else:
                    v = ""
                print(v)
                # 拼接文本
                # strs = "\"{k}\" = \"{v}\";\n".format(k=k, v=v)
                # stringls.append(strs)

            # 拼接字符串
            # text = "".join(stringls)
            # print(text)


    def parse_xlsx_file(self, xlsx_fname):
        wb = load_workbook(xlsx_fname)
        sheets = wb.sheetnames
        for sh in sheets:
            print(sh)  # ['Sheet1', 'Sheet2', 'Sheet3']
        sheet = wb.get_sheet_by_name(sheets[0])
        print(sheet)
        # print(sheet["C"])  # 第C列

        # 获取语言列表
        lans = sheet["1"][1:]  # 第1行
        for l in lans:
            if l.value is not None:
                print(l.column)
                print(l.value)
                self.langls.append(l.value)

        print(sheet["C4"].value)  # c4     <-第C4格的值
        print(sheet.max_row)  # 10     <-最大行数
        print(sheet.max_column)  # 5     <-最大列数

        for i in range(2,sheet.max_row): # <-C列中的所有值
            subls = []
            row_str = sheet[str(i)]
            for r in row_str:
                if r.value is not None:
                    print(r.value)
                    subls.append(r.value.encode("utf-8"))
            if subls:
                self.infols.append({"key": subls[0], "value": subls[1:]})


    # 解析txt文本
    def parse_txt_file(self, txt_fname):
        # 读取文本信息
        cwd = os.path.dirname(sys.argv[0])
        self.txt_fd = codecs.open(os.path.join(cwd, txt_fname), 'r', 'utf-8')

        # 读取每一行的数据
        ls = [line.strip().encode('utf-8') for line in self.txt_fd]
        self.txt_fd.close()

        # 对第一行的数据进行划分,获取语言列表
        self.langls = ls[0].split('\t\t')[1:]

        for i in range(len(self.langls)):
            if self.langls[i] == "中文":
                self.langls[i] = "zh"
            elif self.langls[i] == "英文":
                self.langls[i] = "en"
            elif self.langls[i] == "韩文":
                self.langls[i] = "ko"
            elif self.langls[i] == "日文":
                self.langls[i] = "ja"
            elif self.langls[i] == "俄语":
                self.langls[i] = "ru"
            elif self.langls[i] == "西班牙语":
                self.langls[i] = "es"

        for lang in self.langls:
            print("langls:{}".format(lang))

        null_ls = []
        # 对每一行的数据进行分割,将数据存入list(第一列为键,其他列为值)
        for i in ls[1:]:
            subls = i.split('\t\t')
            if subls[0] != "-":
                self.infols.append({"key": subls[0], "value": subls[1:]})
            else:
                null_ls.append({"key": subls[0], "value": subls[1:]})

                # for i in null_ls:
                #     self.infols.append(i)

    # 生成xml文本(Android)
    def generate_xml(self, lang):
        # 新建xml文档
        doc = Document()
        resources = doc.createElement("resources")
        doc.appendChild(resources)

        # 对应的语言索引位置
        langls_index = self.langls.index(lang)
        for info in self.infols:
            key = info["key"]
            try:
                value = info["value"][langls_index]
                # Android的单引号需要转义处理
                if "\'" in value:
                    value = value.replace("\'", "\\\'")
            except IndexError:  # 防止空列越界
                value = ""

            stringele = doc.createElement("string")  # 添加元素
            stringele.setAttribute("name", key)  # 添加并设置属性
            text_node = doc.createTextNode(value)  # 生成节点字符串
            stringele.appendChild(text_node)  # 添加文本节点
            resources.appendChild(stringele)  # 在xml中添加此行数据

        # 拼接字符串
        uglyXml = doc.toprettyxml(indent='  ')
        text_re = re.compile('>\n\s+([^<>\s].*?)\n\s+</', re.DOTALL)
        prettyXml = text_re.sub('>\g<1></', uglyXml)
        return prettyXml

    # 生成键值对的txt文本(iOS)
    def generate_txt(self, lang):
        stringls = []
        # 对应的语言索引位置
        langls_index = self.langls.index(lang)
        for info in self.infols:
            key = info["key"]
            try:
                value = info["value"][langls_index]
                # iOS的双引号需要转义处理
                if "\"" in value:
                    value = value.replace("\"", "\\\"")
            except IndexError:  # 防止空列越界
                value = ""

            # 拼接文本
            str = "\"{k}\" = \"{v}\";\n".format(k=key, v=value)
            stringls.append(str)

        # 拼接字符串
        text = "".join(stringls)
        return text

    # 建立xml文本
    def build_xml_file(self, form="xml"):
        cwd = os.path.dirname(sys.argv[0])
        lsdir = os.listdir('.')

        # 移除"values-"开头的目录及文件,防止重复写入
        for i in lsdir:
            if i.startswith(TEXT_FILE_DIR_PREFIX):
                shutil.rmtree(i)

        # 生成对应语言的文本
        for lang in self.langls:
            # # 生成根据后缀名生成相应格式的文本
            if form == "xml":
                text = self.generate_xml(lang)
            else:
                text = self.generate_txt(lang)

            # 根据语言,建立相应的文件夹目录
            dir_name = TEXT_FILE_DIR_PREFIX + lang
            os.mkdir(dir_name)

            # 新建文本,并将数据写入
            file = "strings." + form
            self.txtfd = open(os.path.join(dir_name, file), 'w')
            self.txtfd.write(text)
            self.txtfd.close()


if __name__ == '__main__':
    obj = unicodetxt_to_formattxt_obj()
    obj.parse_txt_file(sys.argv[1])
    obj.build_xml_file()

    # file = "/Users/Omega/Downloads/多语言翻译.xlsx"
    # obj.parse_xlsx_file(file)
    # obj.build_xml_file()
