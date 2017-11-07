# coding=UTF-8

import os
import codecs
import shutil
import tkMessageBox
import Tkinter
import tkFileDialog
from Tkinter import *
import xlrd
from openpyxl import load_workbook


FILE_CODING = "utf-8"  # 翻译文本编码格式(默认UTF-8)
ITEM_SEPARATOR = "\t"  # 列表间的分隔符
OUTPUT_FORMAT_IOS = "iOS"
OUTPUT_FORMAT_ANDROID = "Android"


class I18NTranslator:
    # 初始化
    def __init__(self):
        self.langls = []  # 语言的list
        self.infols = []  # 信息的list
        self.cur_dir = os.path.split(os.path.realpath(__file__))[0]  # 脚本当前的路径       

    # 解析txt文本
    def parse_txt_file(self, txt_fname):
        # 读取文本信息
        cwd = os.path.dirname(sys.argv[0])
        txt_fd = codecs.open(os.path.join(cwd, txt_fname), 'r', FILE_CODING)
        # 读取每一行的数据
        ls = [line.strip().encode(FILE_CODING) for line in txt_fd]
        txt_fd.close()

        # 对第一行的数据进行划分,获取语言列表
        self.langls = ls[0].split(ITEM_SEPARATOR)[1:]
        # 修改语言描述
        self.edit_language_desc()

        null_ls = []
        # 对每一行的数据进行分割,将数据存入list(第一列为键,其他列为值)
        for i in ls[1:]:
            subls = i.split(ITEM_SEPARATOR)
            if subls[0] != "-":
                self.infols.append({"key": subls[0], "value": subls[1:]})
            else:
                null_ls.append({"key": subls[0], "value": subls[1:]})

    # 解析xlsx格式的Excel文档（按行读取数据）
    # 其实也可以按列读取数据，但是考虑到错位问题（整列的数据比较多），一旦某一行数据移位了，后面key值映射的数据也全乱了，而且由于是同种语言，很不容易发现错误
    # 按行读取的话，每行数据块比较小，即使发生错位了，也很容易根据语言差异定位到错误的位置，所以这里采用按行读取Excel数据的方式
    def parse_xlsx_file(self, xlsx_fname):
        # 获取翻译文档所在的那张表（默认第一张表）
        wb = load_workbook(xlsx_fname)
        sheets = wb.sheetnames
        sheet = wb.get_sheet_by_name(sheets[0])

        # 获取语言列表
        lang_col = []
        for l in sheet["1"][1:]:  # 第1行
            if l.value is not None:
                lang_col.append(l.column)  # 语言列的索引
                self.langls.append(l.value.encode("utf-8"))
        # 修改语言描述
        self.edit_language_desc()

        # 逐行读取数据（从第二行到最后一列）
        for i in range(2, sheet.max_row):
            subls = []
            row_str = sheet[str(i)]
            if row_str[0].value:  # key不为空才进行处理
                for r in row_str:
                    # 将各国语言列下的数据添加到list中
                    if r.column in lang_col:
                        # 需要保留空数据，防止数据移位
                        if r.value is None:
                            info = ""
                        else:
                            info = r.value
                        subls.append(info.encode("utf-8"))
                self.infols.append({"key": row_str[0].value, "value": subls})

    # 解析xls格式的文档（openpyxl不支持xls文档，这里用的是xlrd库，解析方式跟上面类似，因为接口有区别，所以分开写）
    def parse_xls_file(self, xls_fname):
        # 打开文件
        wb = xlrd.open_workbook(xls_fname)
        # 根据sheet索引或者名称获取sheet内容
        sheet = wb.sheet_by_index(0)  # sheet索引从0开始
        # 获取语言列表
        lang_col = []
        first_row = sheet.row(0)
        for i in range(len(first_row)):  # 第1行
            if i > 0 and first_row[i].value is not None:
                lang_col.append(i)  # 语言列的索引
                self.langls.append(first_row[i].value.encode("utf-8"))
        # 修改语言描述
        self.edit_language_desc()

        # 逐行读取数据（从第二行到最后一行）
        max_row = sheet.nrows
        for i in range(1, max_row):
            subls = []
            row_str = sheet.row(i)
            if row_str[0].value:  # key不为空才进行处理
                for r in range(len(row_str)):
                    if r in lang_col:
                        # 将各国语言列下的数据添加到list中
                        if row_str[r].value is None:
                            info = ""
                        else:
                            info = row_str[r].value
                        subls.append(info.encode("utf-8"))
                self.infols.append({"key": row_str[0].value, "value": subls})

    # 修改语言描述（可根据实际情况自定义）
    def edit_language_desc(self):
        for i in range(len(self.langls)):
            if self.langls[i] == "中文":
                self.langls[i] = "zh"
            elif self.langls[i] == "英文":
                self.langls[i] = "en"
            elif self.langls[i] == "韩文":
                self.langls[i] = "ko"
            elif self.langls[i] == "日文":
                self.langls[i] = "ja"
            elif self.langls[i] == "俄语":
                self.langls[i] = "ru"
            elif self.langls[i] == "西班牙语":
                self.langls[i] = "es"
        for lang in self.langls:
            print("langls:{}".format(lang))

    # 拼接字符生成xml文本(Android)
    def generate_xml(self, lang):
        stringls = []
        # 对应的语言索引位置
        langls_index = self.langls.index(lang)
        for info in self.infols:
            key = info["key"]
            try:
                value = info["value"][langls_index]
                # Android的单引号需要转义处理
                if "\'" in value:
                    value = value.replace("\'", "\\\'")
            except IndexError:  # 防止空列越界
                value = ""

            # 拼接每一行的文本
            str = "<string name=\"{k}\">{v}</string>\n".format(k=key, v=value)
            stringls.append(str)

        # 拼接字符串
        text = "<resources>\n\t" + "\t".join(stringls) + "</resources>\n"
        return text

    # 生成键值对的txt文本(iOS)
    def generate_txt(self, lang):
        stringls = []
        # 对应的语言索引位置
        langls_index = self.langls.index(lang)
        for info in self.infols:
            key = info["key"]
            try:
                value = info["value"][langls_index]
                # iOS的双引号需要转义处理
                if "\"" in value:
                    value = value.replace("\"", "\\\"")
            except IndexError:  # 防止空列越界
                value = ""

            # 拼接每一行的文本
            str = "\"{k}\" = \"{v}\";\n".format(k=key, v=value)
            stringls.append(str)

        # 拼接字符串
        text = "".join(stringls)
        return text

    # 按语言分别生成对应的翻译文本
    def build_translated_file(self, form):
        # 切换到脚本的路径
        os.chdir(self.cur_dir)
        # 建立父文件夹,并切换到该目录下
        if form == OUTPUT_FORMAT_ANDROID:
            dir_path = os.path.join(self.cur_dir, OUTPUT_FORMAT_ANDROID)
        else:
            dir_path = os.path.join(self.cur_dir, OUTPUT_FORMAT_IOS)
        if not os.path.exists(dir_path):
            os.makedirs(dir_path)
        os.chdir(dir_path)

        # 移除子目录下所有的文件, 防止写入时发生冲突
        lsdir = os.listdir('.')
        for i in lsdir:
            try:
                # 删除时进行异常处理，防止崩溃（Mac下会自动生成.DS_Store文件）
                shutil.rmtree(i)
            except Exception as e:
                print(e)

        # 生成对应语言的文本
        for lang in self.langls:
            # # 生成根据后缀名生成相应格式的文本
            if form == OUTPUT_FORMAT_ANDROID:
                text = self.generate_xml(lang)
                dir_name = "values-" + lang
                file = "strings.xml"
            else:
                text = self.generate_txt(lang)
                dir_name = lang + ".lproj"
                file = "Localizable.strings"

            # 根据语言,建立相应的文件夹目录
            if not os.path.exists(dir_name):
                os.mkdir(dir_name)

            # 新建文本,并将数据写入
            txtfd = open(os.path.join(dir_name, file), 'w')
            txtfd.write(text)
            txtfd.close()

# 选择文件的图形界面
class SelectFileDialog(Tkinter.Frame):
    def __init__(self, root):
        self.file_path = ""  # 文件路径
        self.path_text = StringVar()  # 路径显示文本
        # 定义文件操作的相关属性
        self.file_opt = options = {}
        options['parent'] = root
        options['initialdir'] = '.'  # 默认当前目录
        options['initialfile'] = 'example.txt'  # 默认选中的文件
        options['title'] = 'Please select a document'  # 标题
        # 限制可选择的文本类型
        options['filetypes'] = [('All', '*.*'), ('Text', '*.txt'), ('Excel', ('*.xls', '*.xlsx'))]

        # 初始化窗口
        Tkinter.Frame.__init__(self, root)
        root.title("Multinational Translation Generator")

        # 添加一个label、entry、button到fileFrame中
        fileFrame = Frame(root)
        fileFrame.pack()
        label = Label(fileFrame, text="path:")
        tvPath = Entry(fileFrame, textvariable=self.path_text)
        btnGetPath = Button(fileFrame, text="select", command=self.select_file)
        label.grid(row=1, column=1)
        tvPath.grid(row=1, column=2)
        btnGetPath.grid(row=1, column=3)

        # 添加两个多选按钮到formatFrame
        formatFrame = Frame(root)
        formatFrame.pack()
        self.value_ios = IntVar()
        self.value_android = IntVar()
        self.value_ios.set(1)
        self.value_android.set(1)
        tvOutput = Label(formatFrame, text="Output Format：")
        cbIOS = Checkbutton(formatFrame, text=OUTPUT_FORMAT_IOS, variable=self.value_ios)
        cbAndroid = Checkbutton(formatFrame, text=OUTPUT_FORMAT_ANDROID, variable=self.value_android)
        # 使用网格管理器排列按钮
        tvOutput.grid(row=1, column=1)
        cbIOS.grid(row=1, column=2)
        cbAndroid.grid(row=1, column=3)

        # 创建Start按钮
        Tkinter.Button(self, text='Start', command=self.begin_transform).pack()

    # 获取文件路径
    def select_file(self):
        self.file_path = tkFileDialog.askopenfilename(**self.file_opt)
        self.path_text.set(self.file_path)

    # 开始转换
    def begin_transform(self):
        if self.file_path:
            opt_ios = self.value_ios.get() == 1
            opt_android = self.value_android.get() == 1
            if opt_ios or opt_android:
                translator = I18NTranslator()
                if self.file_path.endswith(".txt"):  # txt文本
                    translator.parse_txt_file(self.file_path)
                elif self.file_path.endswith(".xls"):  # xls格式文本
                    translator.parse_xls_file(self.file_path)
                elif self.file_path.endswith(".xlsx"):  # xlsx格式文本
                    translator.parse_xlsx_file(self.file_path)
                # 根据选择的格式进行输出
                if opt_ios:
                    translator.build_translated_file(OUTPUT_FORMAT_IOS)
                    print("iOS Finish")
                if opt_android:
                    translator.build_translated_file(OUTPUT_FORMAT_ANDROID)
                    print("Android Finish")
                # 完成提示
                title = "Message"
                tips_string = "Translation Done"
                tkMessageBox.showinfo(title=title, message=tips_string)
        else:
            title = "Error"
            tips_string = "The file path cannot be empty"
            tkMessageBox.showerror(title=title, message=tips_string)


if __name__ == '__main__':
    root = Tkinter.Tk()
    SelectFileDialog(root).pack()
    root.mainloop()
